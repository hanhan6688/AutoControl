from __future__ import annotations

import re
import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import ImportedTestCase, TestCaseFolder, TestPlanProject


class TestPlanImportError(ValueError):
    pass


@dataclass
class SkippedRow:
    """跳过行的详细信息"""

    row_number: int
    reason: str
    raw_data: dict[str, str] = field(default_factory=dict)


@dataclass
class ImportResult:
    """导入结果"""

    plan: TestPlanProject
    imported_count: int
    skipped_count: int
    skipped_rows: list[SkippedRow] = field(default_factory=list)
    duplicate_count: int = 0

    def to_summary(self) -> dict[str, Any]:
        return {
            "imported": self.imported_count,
            "skipped": self.skipped_count,
            "duplicates": self.duplicate_count,
            "skip_reasons": self._group_skip_reasons(),
        }

    def _group_skip_reasons(self) -> dict[str, int]:
        reasons: dict[str, int] = {}
        for row in self.skipped_rows:
            reasons[row.reason] = reasons.get(row.reason, 0) + 1
        return reasons


class TestPlanImportService:
    # 必填列（带 * 前缀）
    REQUIRED_COLUMNS = {
        "*所属系统": "system_name",
        "*所属模块": "module",
        "*用例名称": "case_name",
        "*前置条件": "precondition",
        "*用例步骤": "steps",
        "*预期结果": "expected_result",
        "*关联需求ID": "requirement_id",
    }

    # 可选列
    OPTIONAL_COLUMNS = {
        "用例类型": "case_type",
        "用例等级": "priority",
        "测试模块": "test_module",
        "目标应用": "target_app",
    }

    COLUMN_ALIASES = {
        "所属系统": "system_name",
        "系统": "system_name",
        "system": "system_name",
        "system_name": "system_name",
        "所属模块": "module",
        "模块": "module",
        "module": "module",
        "用例名称": "case_name",
        "AutoGLM": "case_name",
        "用例": "case_name",
        "case": "case_name",
        "case_name": "case_name",
        "name": "case_name",
        "title": "case_name",
        "前置条件": "precondition",
        "前置": "precondition",
        "precondition": "precondition",
        "用例步骤": "steps",
        "测试步骤": "steps",
        "步骤": "steps",
        "steps": "steps",
        "step": "steps",
        "预期结果": "expected_result",
        "期望结果": "expected_result",
        "预期": "expected_result",
        "expected": "expected_result",
        "expected_result": "expected_result",
        "关联需求ID": "requirement_id",
        "需求ID": "requirement_id",
        "需求编号": "requirement_id",
        "requirement_id": "requirement_id",
        "用例类型": "case_type",
        "类型": "case_type",
        "case_type": "case_type",
        "用例等级": "priority",
        "优先级": "priority",
        "priority": "priority",
        "测试模块": "test_module",
        "test_module": "test_module",
        "目标应用": "target_app",
        "应用": "target_app",
        "target_app": "target_app",
    }

    REQUIRED_COLUMN_LABELS = {
        "case_name": "用例名称",
        "steps": "用例步骤",
        "expected_result": "预期结果",
    }

    # 必填字段（数据库层面必须有值）
    REQUIRED_FIELDS = {"case_name", "steps", "expected_result"}

    # 字段长度限制
    FIELD_MAX_LENGTH = {
        "case_name": 500,
        "system_name": 128,
        "module": 128,
        "precondition": 2000,
        "expected_result": 2000,
        "requirement_id": 255,
        "case_type": 64,
        "priority": 32,
        "target_app": 128,
        "test_module": 128,
    }

    # 步骤数量限制
    MAX_STEPS = 50
    MIN_STEPS = 1
    MAX_STEP_LENGTH = 500

    def __init__(self, db: Session) -> None:
        self.db = db

    def import_file(
        self,
        file_path: Path,
        project_name: str,
        source_filename: str | None = None,
        check_duplicates: bool = True,
    ) -> ImportResult:
        """导入AutoGLM文件，返回详细的导入结果"""
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self.import_csv(
                file_path,
                project_name,
                source_filename=source_filename,
                check_duplicates=check_duplicates,
            )
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return self.import_excel(
                file_path,
                project_name,
                source_filename=source_filename,
                check_duplicates=check_duplicates,
            )
        if suffix == ".xls":
            raise TestPlanImportError("暂不支持旧版 .xls，请另存为 .xlsx 或 .csv 后导入")
        raise TestPlanImportError("仅支持 .xlsx、.xlsm 或 .csv AutoGLM文件")

    def import_excel(
        self,
        file_path: Path,
        project_name: str,
        source_filename: str | None = None,
        check_duplicates: bool = True,
    ) -> ImportResult:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.worksheets[0]
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not rows:
            raise TestPlanImportError("AutoGLM文件为空")

        return self._import_rows(
            rows,
            file_path,
            project_name,
            source_filename=source_filename,
            check_duplicates=check_duplicates,
        )

    def import_csv(
        self,
        file_path: Path,
        project_name: str,
        source_filename: str | None = None,
        check_duplicates: bool = True,
    ) -> ImportResult:
        with file_path.open("r", encoding="utf-8-sig", newline="") as csv_file:
            rows = list(csv.reader(csv_file))
        if not rows:
            raise TestPlanImportError("AutoGLM文件为空")

        return self._import_rows(
            rows,
            file_path,
            project_name,
            source_filename=source_filename,
            check_duplicates=check_duplicates,
        )

    def _import_rows(
        self,
        rows: list[Any],
        file_path: Path,
        project_name: str,
        source_filename: str | None = None,
        check_duplicates: bool = True,
    ) -> ImportResult:
        header = [self._clean_cell(value) for value in rows[0]]
        column_map = self._build_column_map(header)

        plan = TestPlanProject(
            name=project_name.strip() or file_path.stem,
            source_filename=source_filename or file_path.name,
            total_cases=0,
        )
        self.db.add(plan)
        self.db.flush()

        # Auto-create folders by grouping on module/system_name field
        folder_by_group: dict[str, TestCaseFolder] = {}
        folder_sequence = 0

        cases: list[ImportedTestCase] = []
        skipped_rows: list[SkippedRow] = []
        duplicate_count = 0
        seen_case_names: set[str] = set()  # 当前批次内的重复检测

        for row_index, raw_row in enumerate(rows[1:], start=2):  # start=2 因为第1行是表头
            values = [self._clean_cell(value) for value in raw_row]

            # 过滤条件1: 空行
            if not any(values):
                continue

            # 过滤条件2: 说明行（包含"必填"或全是"说明行"）
            if self._is_instruction_row(values):
                skipped_rows.append(SkippedRow(
                    row_number=row_index,
                    reason="说明行",
                    raw_data={"raw": " | ".join(str(v) for v in values[:3])},
                ))
                continue

            row_data = self._row_to_data(values, column_map, header)
            case_name = row_data.get("case_name", "").strip()
            steps_raw = row_data.get("steps", "").strip()
            expected_result = row_data.get("expected_result", "").strip()
            precondition = row_data.get("precondition", "").strip()

            # 过滤条件3: 必填字段为空
            missing_fields = self._check_required_fields(case_name, steps_raw, expected_result, precondition)
            if missing_fields:
                skipped_rows.append(SkippedRow(
                    row_number=row_index,
                    reason=f"必填字段为空: {', '.join(missing_fields)}",
                    raw_data={"case_name": case_name, "steps": steps_raw[:50] + "..." if len(steps_raw) > 50 else steps_raw},
                ))
                continue

            # 过滤条件4: 数据校验
            validation_errors = self._validate_row_data(row_data)
            if validation_errors:
                skipped_rows.append(SkippedRow(
                    row_number=row_index,
                    reason=f"数据校验失败: {'; '.join(validation_errors)}",
                    raw_data={"case_name": case_name},
                ))
                continue

            # 过滤条件5: 重复检测
            if check_duplicates:
                # 检测当前批次内重复
                if case_name in seen_case_names:
                    duplicate_count += 1
                    skipped_rows.append(SkippedRow(
                        row_number=row_index,
                        reason="用例名称在当前文件中重复",
                        raw_data={"case_name": case_name},
                    ))
                    continue
                # 检测数据库中已存在
                if self._check_duplicate_in_db(plan.id, case_name):
                    duplicate_count += 1
                    skipped_rows.append(SkippedRow(
                        row_number=row_index,
                        reason="用例名称在数据库中已存在",
                        raw_data={"case_name": case_name},
                    ))
                    continue
                seen_case_names.add(case_name)

            # 解析步骤
            steps = self._parse_steps(steps_raw)
            if not steps:
                skipped_rows.append(SkippedRow(
                    row_number=row_index,
                    reason="步骤解析失败，无有效步骤",
                    raw_data={"case_name": case_name, "steps": steps_raw[:50]},
                ))
                continue

            # Auto-group by system_name to create folders
            group_key = row_data.get("system_name", "").strip()
            if group_key and group_key not in folder_by_group:
                folder_sequence += 1
                folder = TestCaseFolder(
                    plan_id=plan.id,
                    name=group_key,
                    source_type="import_grouped",
                    source_filename=source_filename or file_path.name,
                    sequence=folder_sequence,
                    total_cases=0,
                )
                self.db.add(folder)
                self.db.flush()
                folder_by_group[group_key] = folder

            case = ImportedTestCase(
                plan_id=plan.id,
                folder_id=folder_by_group[group_key].id if group_key else None,
                sequence=len(cases) + 1,
                system_name=self._truncate_field(row_data.get("system_name"), "system_name"),
                module=self._truncate_field(row_data.get("module"), "module"),
                case_name=case_name,
                precondition=self._truncate_field(row_data.get("precondition"), "precondition"),
                steps=steps,
                expected_result=self._truncate_field(expected_result, "expected_result"),
                requirement_id=self._truncate_field(row_data.get("requirement_id"), "requirement_id"),
                case_type=self._truncate_field(row_data.get("case_type"), "case_type"),
                priority=self._truncate_field(row_data.get("priority"), "priority"),
                target_app=self._truncate_field(row_data.get("target_app"), "target_app"),
                test_module=self._truncate_field(row_data.get("test_module"), "test_module"),
                run_count=0,
                latest_result="pending",
                latest_result_note="",
            )
            self.db.add(case)
            cases.append(case)

        if not cases:
            # 回滚以避免创建空的测试计划
            self.db.rollback()
            summary = self._build_error_summary(skipped_rows, duplicate_count)
            raise TestPlanImportError(f"没有解析到有效AutoGLM。\n{summary}")

        plan.total_cases = len(cases)

        # Update folder total_cases counts
        folder_case_counts: dict[int, int] = {}
        for case in cases:
            if case.folder_id:
                folder_case_counts[case.folder_id] = folder_case_counts.get(case.folder_id, 0) + 1
        for fid, count in folder_case_counts.items():
            folder = folder_by_group.get(
                # Find the group key that maps to this folder
                next(k for k, v in folder_by_group.items() if v.id == fid),
                None,
            )
            if folder:
                folder.total_cases = count

        self.db.commit()
        self.db.refresh(plan)

        return ImportResult(
            plan=plan,
            imported_count=len(cases),
            skipped_count=len(skipped_rows),
            skipped_rows=skipped_rows,
            duplicate_count=duplicate_count,
        )

    def _build_column_map(self, header: list[str]) -> dict[int, str]:
        """构建列索引到字段名的映射"""
        result: dict[int, str] = {}
        all_columns = {**self.REQUIRED_COLUMNS, **self.OPTIONAL_COLUMNS}

        for index, title in enumerate(header):
            field_name = all_columns.get(title) or self.COLUMN_ALIASES.get(self._normalize_header(title))
            if field_name:
                result[index] = field_name

        # 检查必填列是否存在
        present_fields = set(result.values())
        missing = [
            label
            for field_name, label in self.REQUIRED_COLUMN_LABELS.items()
            if field_name not in present_fields
        ]
        if missing:
            raise TestPlanImportError(f"AutoGLM文件缺少必填列: {', '.join(missing)}")

        return result

    @staticmethod
    def _normalize_header(title: str) -> str:
        return re.sub(r"[\s*_＊:：\-—]+", "", str(title or "").strip())

    @staticmethod
    def _row_to_data(values: list[str], column_map: dict[int, str], header: list[str]) -> dict[str, str]:
        """将行数据转换为字典"""
        data: dict[str, str] = {}
        for index, field_name in column_map.items():
            data[field_name] = values[index] if index < len(values) else ""
        return data

    def _check_required_fields(
        self,
        case_name: str,
        steps: str,
        expected_result: str,
        precondition: str = "",
    ) -> list[str]:
        """检查必填字段，返回缺失的字段列表

        Note: precondition（前置条件）is NOT required here because the database
        model allows NULL.  Rows with an empty precondition will still be imported.
        """
        missing = []
        if not case_name:
            missing.append("用例名称")
        if not steps:
            missing.append("用例步骤")
        if not expected_result:
            missing.append("预期结果")
        return missing

    def _validate_row_data(self, row_data: dict[str, str]) -> list[str]:
        """数据校验，返回错误列表"""
        errors: list[str] = []

        # 用例名称长度校验
        case_name = row_data.get("case_name", "")
        if len(case_name) > self.FIELD_MAX_LENGTH.get("case_name", 500):
            errors.append(f"用例名称超过{self.FIELD_MAX_LENGTH['case_name']}字符")

        # 步骤数量校验（预解析）
        steps_raw = row_data.get("steps", "")
        step_count = len(self._parse_steps(steps_raw))
        if step_count > self.MAX_STEPS:
            errors.append(f"步骤数量超过{self.MAX_STEPS}条")
        elif step_count < self.MIN_STEPS:
            errors.append("至少需要1个步骤")

        # 特殊字符校验（防止注入）
        dangerous_patterns = ["<script", "javascript:", "onerror=", "onload="]
        for field in ["case_name", "expected_result", "precondition"]:
            value = row_data.get(field, "").lower()
            for pattern in dangerous_patterns:
                if pattern in value:
                    errors.append(f"{field}包含不允许的内容")
                    break

        return errors

    def _check_duplicate_in_db(self, plan_id: int, case_name: str) -> bool:
        """检查数据库中是否已存在相同用例名称"""
        existing = self.db.query(ImportedTestCase).filter(
            ImportedTestCase.plan_id == plan_id,
            ImportedTestCase.case_name == case_name,
        ).first()
        return existing is not None

    def _truncate_field(self, value: str | None, field_name: str) -> str | None:
        """截断超长字段"""
        if not value:
            return None
        max_length = self.FIELD_MAX_LENGTH.get(field_name)
        if max_length and len(value) > max_length:
            return value[:max_length - 3] + "..."
        return value

    @staticmethod
    def _build_error_summary(skipped_rows: list[SkippedRow], duplicate_count: int) -> str:
        """构建错误摘要"""
        if not skipped_rows:
            return ""

        reasons: dict[str, int] = {}
        for row in skipped_rows:
            reasons[row.reason] = reasons.get(row.reason, 0) + 1

        lines = ["导入统计:"]
        for reason, count in sorted(reasons.items(), key=lambda x: -x[1]):
            lines.append(f"  - {reason}: {count}行")

        if duplicate_count > 0:
            lines.append(f"  - 重复用例: {duplicate_count}条")

        return "\n".join(lines)

    @staticmethod
    def _parse_steps(raw_steps: str) -> list[str]:
        steps: list[str] = []
        for part in re.split(r"(?:\r?\n)+|(?:^|\s)(?=\d+[.、])", raw_steps):
            value = re.sub(r"^\s*\d+[.、]\s*", "", part).strip()
            if value:
                steps.append(value)
        return steps

    @staticmethod
    def _is_instruction_row(values: list[str]) -> bool:
        joined = " ".join(values)
        return "必填" in joined or all(value == "说明行" for value in values if value)

    @staticmethod
    def _clean_cell(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()
