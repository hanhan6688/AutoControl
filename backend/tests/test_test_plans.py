from pathlib import Path
import io
import json

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session, sessionmaker

from app.database import Base


def build_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AI学区顾问"
    sheet.append(
        [
            "*所属系统",
            "*所属模块",
            "*用例名称",
            "*前置条件",
            "*用例步骤",
            "*预期结果",
            "*关联需求ID",
            "用例类型",
            "用例等级",
        ]
    )
    sheet.append(["说明行", "说明行", "说明行", "说明行", "说明行", "说明行", "说明行", "说明行", "说明行"])
    sheet.append(
        [
            "AI找房",
            "AI找房",
            "验证点击入口跳转",
            "用户在首页",
            "1.点击AI学区顾问入口\n2.等待页面加载",
            "跳转到AI学区顾问页面",
            "46546",
            "功能测试",
            "高",
        ]
    )
    sheet.append(
        [
            "AI找房",
            "AI找房",
            "验证初始页面显示输入框",
            "用户在AI学区顾问页面",
            "1.查看页面底部\n2.查看是否有输入框",
            "页面底部显示输入框",
            "46546",
            "功能测试",
            "中",
        ]
    )
    workbook.save(path)


def build_csv(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "*所属系统,*所属模块,*用例名称,*前置条件,*用例步骤,*预期结果,*关联需求ID,用例类型,用例等级",
                "说明行,说明行,说明行,说明行,说明行,说明行,说明行,说明行,说明行",
                "AI找房,AI找房,验证点击入口跳转,用户在首页,\"1.点击AI学区顾问入口\n2.等待页面加载\",跳转到AI学区顾问页面,46546,功能测试,高",
                "AI找房,AI找房,验证初始页面显示输入框,用户在AI学区顾问页面,\"1.查看页面底部\n2.查看是否有输入框\",页面底部显示输入框,46546,功能测试,中",
            ]
        ),
        encoding="utf-8-sig",
    )


def make_session() -> Session:
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    return sessionmaker(bind=engine, autoflush=False, autocommit=False)()


def test_import_service_expands_excel_rows_into_plan_cases(tmp_path: Path) -> None:
    from app.services.test_plan_import_service import TestPlanImportService

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)
    db = make_session()

    result = TestPlanImportService(db).import_excel(workbook_path, project_name="AI学区顾问搜索交互")

    assert result.plan.name == "AI学区顾问搜索交互"
    assert result.plan.total_cases == 2
    assert result.imported_count == 2
    assert result.skipped_count == 1  # 说明行被跳过
    assert [case.sequence for case in result.plan.cases] == [1, 2]
    assert result.plan.cases[0].case_name == "验证点击入口跳转"
    assert result.plan.cases[0].steps == ["点击AI学区顾问入口", "等待页面加载"]
    assert result.plan.cases[0].expected_result == "跳转到AI学区顾问页面"
    assert result.plan.cases[0].run_count == 0
    assert result.plan.cases[0].latest_result == "pending"


def test_import_service_expands_csv_rows_into_plan_cases(tmp_path: Path) -> None:
    from app.services.test_plan_import_service import TestPlanImportService

    csv_path = tmp_path / "cases.csv"
    build_csv(csv_path)
    db = make_session()

    result = TestPlanImportService(db).import_file(csv_path, project_name="AI学区顾问搜索交互")

    assert result.plan.source_filename == "cases.csv"
    assert result.plan.total_cases == 2
    assert result.imported_count == 2
    assert result.plan.cases[0].case_name == "验证点击入口跳转"
    assert result.plan.cases[0].steps == ["点击AI学区顾问入口", "等待页面加载"]
    assert result.plan.cases[1].expected_result == "页面底部显示输入框"


def test_import_service_accepts_simple_csv_headers(tmp_path: Path) -> None:
    from app.services.test_plan_import_service import TestPlanImportService

    csv_path = tmp_path / "simple_cases.csv"
    csv_path.write_text(
        "\n".join(
            [
                "用例名称,步骤,预期结果,目标应用,测试模块",
                "启动AI找房,\"1.启动乐有家\n2.进入AI找房页面\",展示AI找房页面,乐有家测试版,AI找房",
            ]
        ),
        encoding="utf-8-sig",
    )
    db = make_session()

    result = TestPlanImportService(db).import_file(csv_path, project_name="简化CSV")

    assert result.imported_count == 1
    assert result.plan.cases[0].case_name == "启动AI找房"
    assert result.plan.cases[0].steps == ["启动乐有家", "进入AI找房页面"]
    assert result.plan.cases[0].target_app == "乐有家测试版"
    assert result.plan.cases[0].test_module == "AI找房"


def test_import_endpoint_returns_400_for_invalid_csv_headers(tmp_path: Path) -> None:
    from app import database
    from app.main import create_app

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    database.engine = engine
    database.SessionLocal.configure(bind=engine)
    Base.metadata.create_all(bind=engine)

    csv_path = tmp_path / "invalid.csv"
    csv_path.write_text("标题,内容\n只有标题,没有步骤列\n", encoding="utf-8-sig")

    client = TestClient(create_app(), raise_server_exceptions=False)
    with csv_path.open("rb") as file:
        response = client.post(
            "/api/test-plans/import",
            data={"project_name": "错误CSV"},
            files={"file": ("invalid.csv", file, "text/csv")},
        )

    assert response.status_code == 400
    assert "缺少必填列" in response.json()["detail"]
    assert "用例名称" in response.json()["detail"]


def test_execution_service_creates_run_and_updates_case_report(tmp_path: Path) -> None:
    from app.services.test_execution_service import TestExecutionService
    from app.services.test_plan_import_service import TestPlanImportService

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)
    db = make_session()
    result = TestPlanImportService(db).import_excel(workbook_path, project_name="AI学区顾问搜索交互")

    run = TestExecutionService(db).execute_case(result.plan.cases[0].id, device_udid=None)
    db.refresh(result.plan.cases[0])

    assert run.case_id == result.plan.cases[0].id
    assert run.run_result == "failed"
    assert "未选择设备" in run.result_note
    assert run.error_category == "no_device"
    assert result.plan.cases[0].run_count == 1
    assert result.plan.cases[0].latest_result == "failed"
    assert result.plan.cases[0].latest_result_note == run.result_note


def test_execution_service_saves_into_legacy_autoglm_configured_schema(tmp_path: Path) -> None:
    from app.services.test_execution_service import TestExecutionService
    from app.services.test_plan_import_service import TestPlanImportService

    db_path = tmp_path / "legacy.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    with engine.begin() as connection:
        columns = {
            row[1]
            for row in connection.exec_driver_sql("PRAGMA table_info(test_case_execution)")
        }
        if "autoglm_configured" not in columns:
            connection.exec_driver_sql(
                "ALTER TABLE test_case_execution ADD COLUMN autoglm_configured BOOLEAN NOT NULL"
            )

    db = sessionmaker(bind=engine, autoflush=False, autocommit=False)()
    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)
    result = TestPlanImportService(db).import_excel(workbook_path, project_name="AI学区顾问搜索交互")

    run = TestExecutionService(db).execute_case(result.plan.cases[0].id, device_udid=None)

    stored_value = db.execute(
        text("SELECT autoglm_configured FROM test_case_execution WHERE id = :id"),
        {"id": run.id},
    ).scalar_one()
    assert run.run_result == "failed"
    assert "未选择设备" in run.result_note
    assert stored_value in {0, False}


def test_execution_service_maps_three_mobile_platforms_to_autoglm_device_types() -> None:
    from app.services.test_execution_service import TestExecutionService

    service = TestExecutionService(make_session())

    android_command = service._build_autoglm_command("android", "android-1", "task")
    ios_command = service._build_autoglm_command("ios", "ios-1", "task")
    harmony_command = service._build_autoglm_command("harmony", "harmony-1", "task")

    assert android_command[android_command.index("--device-type") + 1] == "adb"
    assert ios_command[ios_command.index("--device-type") + 1] == "ios"
    assert "--wda-url" in ios_command
    assert harmony_command[harmony_command.index("--device-type") + 1] == "hdc"


def test_execution_service_sets_wait_takeover_mode_for_stream_runs(monkeypatch) -> None:
    from app.services.test_execution_service import TestExecutionService

    monkeypatch.delenv("PHONE_AGENT_TAKEOVER_MODE", raising=False)
    service = TestExecutionService(make_session())

    stream_env = service._autoglm_env("android", "android-1", takeover_mode="wait")
    non_stream_env = service._autoglm_env("android", "android-1", takeover_mode="fail")

    assert stream_env["PHONE_AGENT_TAKEOVER_MODE"] == "wait"
    assert non_stream_env["PHONE_AGENT_TAKEOVER_MODE"] == "fail"


def test_autoglm_phase_writes_step_screenshots_and_terminal_log(tmp_path: Path, monkeypatch) -> None:
    from app.services.test_execution_service import TestExecutionService

    service = TestExecutionService(make_session())
    monkeypatch.setattr(service, "_build_autoglm_command", lambda platform, device_udid, prompt, screenshot_dir=None: ["autoglm"])

    # Pre-create screenshots that AutoGLM would have saved via STEP_SCREENSHOT_SAVED
    screenshots_dir = tmp_path / "screenshots"
    screenshots_dir.mkdir(parents=True, exist_ok=True)
    (screenshots_dir / "step_001.png").write_bytes(b"png")
    (screenshots_dir / "step_002.png").write_bytes(b"png")

    class FakeProcess:
        pid = 1234

        def __init__(self) -> None:
            # Real stdout order: 🎯 → JSON → ===== → STEP_SCREENSHOT_SAVED
            self.stdout = iter([
                "==================================================\n",
                "💭 思考过程:\n",
                "--------------------------------------------------\n",
                "thinking text\n",
                "🎯 执行动作:\n",
                '{"_metadata": "do", "action": "Tap", "coordinate": [100, 200]}\n',
                "==================================================\n",
                "STEP_SCREENSHOT_SAVED:step_001.png\n",
                "==================================================\n",
                "💭 思考过程:\n",
                "--------------------------------------------------\n",
                "thinking text\n",
                "🎯 执行动作:\n",
                '{"_metadata": "finish", "message": "done"}\n',
                "==================================================\n",
                "STEP_SCREENSHOT_SAVED:step_002.png\n",
            ])
            self.stderr = io.StringIO("diagnostic warning\n")

        def poll(self):
            return None

        def kill(self) -> None:
            return None

        def wait(self) -> int:
            return 0

    monkeypatch.setattr("app.services.test_execution_service.subprocess.Popen", lambda *args, **kwargs: FakeProcess())

    events = list(service._run_autoglm_phase(
        phase="execution",
        prompt="task",
        platform="android",
        device_udid="device-1",
        report_dir=tmp_path,
        emit=lambda event_type, phase, message, **extra: {
            "event": event_type,
            "type": extra.pop("type", event_type),
            "phase": phase,
            "message": message,
            **extra,
        },
    ))

    # Screenshots are now merged into action_executed events
    action_events = [event for event in events if event.get("type") == "action_executed"]
    terminal_event = next(event for event in events if event.get("type") == "autoglm_terminal_log")
    terminal_path = tmp_path / terminal_event["terminal_log_url"].split("/")[-1]

    assert len(action_events) == 2
    assert action_events[0].get("screenshot_url") is not None
    assert action_events[1].get("screenshot_url") is not None
    assert (tmp_path / "screenshots" / "step_001.png").exists()
    assert (tmp_path / "screenshots" / "step_002.png").exists()
    assert terminal_path.exists()
    terminal_text = terminal_path.read_text(encoding="utf-8")
    assert "[stdout] STEP_SCREENSHOT_SAVED:step_001.png" in terminal_text
    assert "[stderr] diagnostic warning" in terminal_text


def test_execution_service_saves_orchestrated_task_plan(tmp_path: Path, monkeypatch) -> None:
    from app.config import settings
    from app.services.test_execution_service import TestExecutionService
    from app.services.test_plan_import_service import TestPlanImportService

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)
    db = make_session()
    result = TestPlanImportService(db).import_excel(workbook_path, project_name="AI学区顾问搜索交互")

    run = TestExecutionService(db).execute_case(result.plan.cases[0].id, device_udid=None)

    plan_events = [event for event in run.action_trace if event.get("type") == "case_task_plan"]
    assert plan_events
    # After CaseOrchestrationService removal, emit just carries target_kind
    assert plan_events[0]["target_kind"] is not None


def test_result_assertion_fallback_passes_when_logs_match_expectation(monkeypatch) -> None:
    from app.config import settings
    from app.models import ImportedTestCase
    from app.services.result_assertion_service import ResultAssertionService

    monkeypatch.setattr(settings, "result_assertion_enabled", False)
    case = ImportedTestCase(
        plan_id=1,
        sequence=1,
        case_name="验证跳转结果",
        precondition="用户在首页",
        steps=["点击AI学区顾问入口", "等待页面加载"],
        expected_result="跳转到AI学区顾问页面",
        target_app="乐有家测试版",
        test_module="AI找房",
    )
    logs = [
        {"phase": "execution", "message": "点击AI学区顾问入口"},
        {"phase": "execution", "message": "页面已跳转到AI学区顾问页面"},
    ]

    assertion = ResultAssertionService().assert_result(case=case, action_trace=logs)

    assert assertion["verdict"] == "passed"
    assert assertion["source"] == "fallback"
    assert assertion["confidence"] >= 0.6


def test_result_assertion_uses_screenshot_as_visual_evidence(tmp_path: Path, monkeypatch) -> None:
    from app.config import settings
    from app.models import ImportedTestCase
    from app.services import result_assertion_service
    from app.services.result_assertion_service import ResultAssertionService

    monkeypatch.setattr(settings, "result_assertion_enabled", True)
    monkeypatch.setattr(settings, "result_assertion_api_key", "test-key")
    monkeypatch.setattr(settings, "result_assertion_base_url", "https://example.com/v1")
    monkeypatch.setattr(settings, "result_assertion_model", "vision-model")

    screenshot_path = tmp_path / "final.png"
    screenshot_path.write_bytes(b"\x89PNG\r\n\x1a\nfake")
    captured_messages = []

    class FakeMessage:
        content = json.dumps({
            "verdict": "passed",
            "confidence": 0.91,
            "reason": "截图显示目标页面已经打开。",
            "evidence": ["终态截图包含目标页面标题"],
            "failed_expectations": [],
        }, ensure_ascii=False)

    class FakeChoice:
        message = FakeMessage()

    class FakeCompletions:
        def create(self, **kwargs):
            captured_messages.extend(kwargs["messages"])
            return type("Response", (), {"choices": [FakeChoice()]})()

    class FakeOpenAI:
        def __init__(self, **kwargs):
            pass

        chat = type("Chat", (), {"completions": FakeCompletions()})()

    monkeypatch.setattr(result_assertion_service, "OpenAI", FakeOpenAI)
    case = ImportedTestCase(
        plan_id=1,
        sequence=1,
        case_name="验证AI找房页面",
        precondition="用户在首页",
        steps=["点击AI找房"],
        expected_result="进入AI找房页面",
        target_app="乐有家测试版",
        test_module="AI找房",
    )

    assertion = ResultAssertionService().assert_result(
        case=case,
        action_trace=[{"phase": "execution", "message": "AutoGLM执行完成"}],
        final_screenshot={"url": "/static/reports/final.png", "path": str(screenshot_path)},
    )

    assert assertion["source"] == "ai"
    assert assertion["verdict"] == "passed"
    user_content = captured_messages[1]["content"]
    assert isinstance(user_content, list)
    assert any(part.get("type") == "image_url" for part in user_content)


def test_uncertain_assertion_marks_execution_for_review(tmp_path: Path, monkeypatch) -> None:
    from app.config import settings
    from app.services.test_execution_service import TestExecutionService
    from app.services.test_plan_import_service import TestPlanImportService
    monkeypatch.setattr(settings, "result_assertion_enabled", False)
    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)
    db = make_session()
    result = TestPlanImportService(db).import_excel(workbook_path, project_name="AI学区顾问搜索交互")

    service = TestExecutionService(db)
    monkeypatch.setattr(service, "_resolve_device", lambda device_udid, platform: ("android", None, []))
    monkeypatch.setattr(service, "_validate_open_autoglm_root", lambda: None)
    monkeypatch.setattr(service, "_capture_final_screenshot", lambda report_dir, platform, device_udid: {
        "url": "/static/reports/runs/case/final.png",
        "path": str(tmp_path / "final.png"),
        "created_at": "2026-05-11T00:00:00.000Z",
    })

    def fake_run_autoglm_phase(phase, prompt, platform, device_udid, report_dir, emit, client_run_id=None, trace_id=None, login_accounts=None):
        yield emit("log", phase, "AutoGLM执行完成，但日志没有出现预期关键词")
        return 0

    monkeypatch.setattr(service, "_run_autoglm_phase", fake_run_autoglm_phase)

    run = service.execute_case(result.plan.cases[0].id, device_udid="device-1", device_platform="android")
    assertion_events = [event for event in run.action_trace if event.get("type") == "result_assertion"]

    assert run.run_result == "uncertain"
    assert run.error_category == "assertion_uncertain"
    assert assertion_events[0]["assertion"]["review_required"] is True
    assert assertion_events[0]["assertion"]["review_status"] == "pending"


def test_report_writer_execution_html_includes_assertion_and_task_plan() -> None:
    from app.services.report_writer import ReportWriter

    html = ReportWriter()._generate_execution_html({
        "execution_id": 12,
        "case_name": "验证AI找房页面",
        "final_result": "uncertain",
        "duration_ms": 1200,
        "result_note": "断言不确定：需要复核",
        "case_task_plan": {
            "source": "ai",
            "unified_goal": "启动乐有家并进入AI找房页面",
            "success_criteria": ["页面展示AI找房输入框"],
            "guardrails": ["不要登录敏感账号"],
        },
        "assertion_result": {
            "source": "ai",
            "verdict": "uncertain",
            "confidence": 0.52,
            "reason": "截图能看到页面，但无法确认输入框是否可用。",
            "evidence": ["终态截图已保存"],
            "failed_expectations": ["未找到输入框可用证据"],
            "review_required": True,
            "review_status": "pending",
        },
        "phases": {},
        "final_screenshot_url": "/static/reports/final.png",
    })

    assert "AI断言结果" in html
    assert "uncertain" in html
    assert "52%" in html
    assert "截图能看到页面" in html
    assert "任务书" in html
    assert "启动乐有家并进入AI找房页面" in html
    assert "/static/reports/final.png" in html


def test_plan_report_counts_uncertain_and_review_required(tmp_path: Path) -> None:
    from app import database
    from app.main import create_app
    from app.models import TestCaseExecution
    from app.services.test_plan_import_service import TestPlanImportService

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    database.engine = engine
    database.SessionLocal.configure(bind=engine)
    Base.metadata.create_all(bind=engine)

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)
    db = database.SessionLocal()
    imported = TestPlanImportService(db).import_excel(workbook_path, project_name="AI学区顾问搜索交互")
    case = imported.plan.cases[0]
    execution = TestCaseExecution(
        plan_id=imported.plan.id,
        case_id=case.id,
        run_index=1,
        device_udid="device-1",
        run_result="uncertain",
        result_note="断言不确定：需要人工复核",
        error_category="assertion_uncertain",
        action_trace=[{
            "type": "result_assertion",
            "assertion": {
                "verdict": "uncertain",
                "confidence": 0.45,
                "review_required": True,
                "review_status": "pending",
            },
        }],
    )
    db.add(execution)
    case.run_count = 1
    case.latest_result = "uncertain"
    case.latest_result_note = execution.result_note
    db.commit()
    plan_id = imported.plan.id
    db.close()

    client = TestClient(create_app())
    response = client.get(f"/api/test-plans/{plan_id}/report")

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["uncertain"] == 1
    assert payload["summary"]["review_required"] == 1
    assert payload["cases"][0]["latest_assertion"]["verdict"] == "uncertain"


def test_execution_service_final_result_uses_assertion_verdict(tmp_path: Path, monkeypatch) -> None:
    from app.config import settings
    from app.services.test_execution_service import TestExecutionService
    from app.services.test_plan_import_service import TestPlanImportService
    monkeypatch.setattr(settings, "result_assertion_enabled", False)
    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)
    db = make_session()
    result = TestPlanImportService(db).import_excel(workbook_path, project_name="AI学区顾问搜索交互")

    service = TestExecutionService(db)
    monkeypatch.setattr(service, "_resolve_device", lambda device_udid, platform: ("android", None, []))
    monkeypatch.setattr(service, "_validate_open_autoglm_root", lambda: None)
    monkeypatch.setattr(service, "_capture_final_screenshot", lambda report_dir, platform, device_udid: {
        "url": "/static/reports/runs/case/final.png",
        "path": str(tmp_path / "final.png"),
        "created_at": "2026-05-11T00:00:00.000Z",
    })

    def fake_run_autoglm_phase(phase, prompt, platform, device_udid, report_dir, emit, client_run_id=None, trace_id=None, login_accounts=None):
        message = "页面已跳转到AI学区顾问页面" if phase == "execution" else "已进入首页"
        yield emit("log", phase, message)
        return 0

    monkeypatch.setattr(service, "_run_autoglm_phase", fake_run_autoglm_phase)

    run = service.execute_case(result.plan.cases[0].id, device_udid="device-1", device_platform="android")

    assertion_events = [event for event in run.action_trace if event.get("type") == "result_assertion"]
    assert run.run_result == "passed"
    assert assertion_events
    assert assertion_events[0]["assertion"]["verdict"] == "passed"


def test_execution_service_externalizes_full_trace_and_keeps_db_trace_slim(tmp_path: Path, monkeypatch) -> None:
    from app.config import settings
    from app.services.test_execution_service import TestExecutionService
    from app.services.test_plan_import_service import TestPlanImportService
    monkeypatch.setattr(settings, "result_assertion_enabled", False)
    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)
    db = make_session()
    result = TestPlanImportService(db).import_excel(workbook_path, project_name="AI学区顾问搜索交互")

    service = TestExecutionService(db)
    monkeypatch.setattr(service, "_resolve_device", lambda device_udid, platform: ("android", None, []))
    monkeypatch.setattr(service, "_validate_open_autoglm_root", lambda: None)
    monkeypatch.setattr(service, "_capture_final_screenshot", lambda report_dir, platform, device_udid: {
        "url": "/static/reports/runs/case/final.png",
        "path": str(tmp_path / "final.png"),
        "created_at": "2026-05-11T00:00:00.000Z",
    })

    def noisy_run_autoglm_phase(phase, prompt, platform, device_udid, report_dir, emit, client_run_id=None, trace_id=None, login_accounts=None):
        for index in range(120):
            message = f"{phase} verbose log {index} " + ("x" * 400)
            yield emit("log", phase, message, stream="stdout")
        if phase == "execution":
            yield emit("log", phase, "页面已跳转到AI学区顾问页面")
        return 0

    monkeypatch.setattr(service, "_run_autoglm_phase", noisy_run_autoglm_phase)

    run = service.execute_case(result.plan.cases[0].id, device_udid="device-1", device_platform="android")
    external_log = next(event for event in run.action_trace if event.get("type") == "external_log")
    log_path = settings.static_dir / external_log["log_url"].removeprefix("/static/")

    assert log_path.exists()
    assert external_log["log_event_count"] > len(run.action_trace)
    assert len(json.dumps(run.action_trace, ensure_ascii=False)) < log_path.stat().st_size // 3
    assert not any(event.get("stream") == "stdout" for event in run.action_trace)


def test_execution_service_writes_log_and_report_once(tmp_path: Path, monkeypatch) -> None:
    from app.config import settings
    from app.services.test_execution_service import TestExecutionService
    from app.services.test_plan_import_service import TestPlanImportService
    monkeypatch.setattr(settings, "result_assertion_enabled", False)
    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)
    db = make_session()
    result = TestPlanImportService(db).import_excel(workbook_path, project_name="AI学区顾问搜索交互")

    service = TestExecutionService(db)
    monkeypatch.setattr(service, "_resolve_device", lambda device_udid, platform: ("android", None, []))
    monkeypatch.setattr(service, "_validate_open_autoglm_root", lambda: None)
    monkeypatch.setattr(service, "_capture_final_screenshot", lambda report_dir, platform, device_udid: {
        "url": "/static/reports/runs/case/final.png",
        "path": str(tmp_path / "final.png"),
        "created_at": "2026-05-11T00:00:00.000Z",
    })

    def fake_run_autoglm_phase(phase, prompt, platform, device_udid, report_dir, emit, client_run_id=None, trace_id=None, login_accounts=None):
        yield emit("log", phase, "页面已跳转到AI学区顾问页面")
        return 0

    monkeypatch.setattr(service, "_run_autoglm_phase", fake_run_autoglm_phase)

    log_write_count = 0
    report_write_count = 0
    original_save_log = service._save_execution_log_file
    original_save_report = service._save_execution_report

    def counting_save_log(*args, **kwargs):
        nonlocal log_write_count
        log_write_count += 1
        return original_save_log(*args, **kwargs)

    def counting_save_report(*args, **kwargs):
        nonlocal report_write_count
        report_write_count += 1
        return original_save_report(*args, **kwargs)

    monkeypatch.setattr(service, "_save_execution_log_file", counting_save_log)
    monkeypatch.setattr(service, "_save_execution_report", counting_save_report)

    run = service.execute_case(result.plan.cases[0].id, device_udid="device-1", device_platform="android")

    assert run.run_result == "passed"
    assert log_write_count == 1
    assert report_write_count == 1


def test_web_execution_emits_pc_agent_model_log(tmp_path: Path, monkeypatch) -> None:
    from app.config import settings
    from app.models import ImportedTestCase, TestPlanProject
    from app.services.test_execution_service import TestExecutionService
    monkeypatch.setattr(settings, "result_assertion_enabled", False)
    monkeypatch.setattr(settings, "pc_agent_provider", "custom_openai")
    monkeypatch.setattr(settings, "pc_agent_model", "local-pc-model")

    db = make_session()
    plan = TestPlanProject(name="Web Plan", total_cases=1)
    db.add(plan)
    db.commit()
    case = ImportedTestCase(
        plan_id=plan.id,
        sequence=1,
        case_name="检查百度首页",
        precondition="打开 https://www.baidu.com",
        steps=["检查页面标题"],
        expected_result="页面显示百度",
        target_app="乐有家测试版",
    )
    db.add(case)
    db.commit()

    class FakeBrowserService:
        opened_sessions = []

        def open(self, url, session=None, headed=False):
            self.opened_sessions.append(session)
            return type("Session", (), {"session_id": session or "fake", "title": "百度", "url": url})()

        def close(self, session=None):
            return None

        def screenshot(self, path, session=None):
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            Path(path).write_bytes(b"png")
            return Path(path)

    class FakeAgentService:
        def __init__(self, browser=None, artifact_root=None):
            pass

        def iter_task_events(self, task, session=None, max_steps=8):
            assert session and session.startswith("web-")
            yield {"event": "result", "phase": "result", "run_result": "passed", "message": "完成"}

    monkeypatch.setattr("app.services.test_execution_service.PCBrowserService", lambda: FakeBrowserService())
    monkeypatch.setattr("app.services.test_execution_service.PCBrowserAgentService", FakeAgentService)
    monkeypatch.setattr(TestExecutionService, "_capture_web_final_screenshot", lambda self, report_dir, browser_service, session=None: {
        "url": "/static/reports/web/final.png",
        "path": str(tmp_path / "final.png"),
        "created_at": "2026-05-21T00:00:00.000Z",
    })

    run = TestExecutionService(db).execute_case(case.id)

    model_events = [event for event in run.action_trace if event.get("type") == "pc_agent_model"]
    assert model_events
    assert model_events[0]["provider"] == "custom_openai"
    assert model_events[0]["model"] == "local-pc-model"


def test_leyoujia_web_execution_requires_saved_auth_state(tmp_path: Path, monkeypatch) -> None:
    from app.config import settings
    from app.models import ImportedTestCase, TestPlanProject
    from app.services.test_execution_service import TestExecutionService
    monkeypatch.setattr(settings, "result_assertion_enabled", False)
    monkeypatch.setattr(TestExecutionService, "_leyoujia_auth_state_path", staticmethod(lambda profile=None: tmp_path / "missing-auth.json"))

    db = make_session()
    plan = TestPlanProject(name="Leyoujia Web Plan", total_cases=1)
    db.add(plan)
    db.commit()
    case = ImportedTestCase(
        plan_id=plan.id,
        sequence=1,
        case_name="检查 zero 测试环境",
        precondition="打开 https://zero-ai-test.leyoujia.com/",
        steps=["检查页面"],
        expected_result="页面可访问",
        target_app="乐有家测试版",
    )
    db.add(case)
    db.commit()

    class FakeBrowserService:
        def open(self, url, session=None, headed=False):
            raise AssertionError("protected Leyoujia target should not open without saved auth state")

    monkeypatch.setattr("app.services.test_execution_service.PCBrowserService", lambda: FakeBrowserService())

    run = TestExecutionService(db).execute_case(case.id)

    assert run.run_result == "failed"
    assert any(event.get("type") == "manual_auth_required" for event in run.action_trace)


def test_leyoujia_prod_web_execution_requires_prod_auth_state(tmp_path: Path, monkeypatch) -> None:
    from app.config import settings
    from app.models import ImportedTestCase, TestPlanProject
    from app.services.test_execution_service import TestExecutionService
    monkeypatch.setattr(settings, "result_assertion_enabled", False)
    monkeypatch.setattr(TestExecutionService, "_leyoujia_auth_state_path", staticmethod(lambda profile=None: tmp_path / "missing-prod-auth.json"))

    db = make_session()
    plan = TestPlanProject(name="Leyoujia Prod Web Plan", total_cases=1)
    db.add(plan)
    db.commit()
    case = ImportedTestCase(
        plan_id=plan.id,
        sequence=1,
        case_name="检查 zero 生产环境",
        precondition="打开 https://zero-ai.leyoujia.com/",
        steps=["检查页面"],
        expected_result="页面可访问",
        target_app="乐有家测试版",
    )
    db.add(case)
    db.commit()

    class FakeBrowserService:
        def open(self, url, session=None, headed=False):
            raise AssertionError("protected Leyoujia prod target should not open without saved auth state")

    monkeypatch.setattr("app.services.test_execution_service.PCBrowserService", lambda: FakeBrowserService())

    run = TestExecutionService(db).execute_case(case.id)

    auth_events = [event for event in run.action_trace if event.get("type") == "manual_auth_required"]
    assert run.run_result == "failed"
    assert auth_events
    assert auth_events[0]["auth_env"] == "prod"
    assert auth_events[0]["login_url"] == "https://i.leyoujia.com/jjslogin/index"


def test_run_stream_endpoint_returns_no_device_event(tmp_path: Path) -> None:
    from app import database
    from app.main import create_app

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    database.engine = engine
    database.SessionLocal.configure(bind=engine)
    Base.metadata.create_all(bind=engine)

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)

    client = TestClient(create_app())
    with workbook_path.open("rb") as file:
        import_response = client.post(
            "/api/test-plans/import",
            data={"project_name": "AI学区顾问搜索交互"},
            files={"file": ("cases.xlsx", file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    case_id = import_response.json()["cases"][0]["id"]
    with client.stream(
        "POST",
        f"/api/test-plans/cases/{case_id}/run/stream",
        json={"device_udid": None, "device_platform": None},
    ) as response:
        assert response.status_code == 200
        body = "".join(response.iter_text())

    assert '"phase": "device_check"' in body
    assert "未选择设备" in body
    assert '"event": "result"' in body


def test_plan_stream_endpoint_groups_case_reports_in_one_folder(tmp_path: Path) -> None:
    from app import database
    from app.main import create_app

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    database.engine = engine
    database.SessionLocal.configure(bind=engine)
    Base.metadata.create_all(bind=engine)

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)

    client = TestClient(create_app())
    with workbook_path.open("rb") as file:
        import_response = client.post(
            "/api/test-plans/import",
            data={"project_name": "AI学区顾问搜索交互"},
            files={"file": ("cases.xlsx", file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    plan_id = import_response.json()["id"]
    with client.stream(
        "POST",
        f"/api/test-plans/{plan_id}/run/stream",
        json={"device_udid": None, "device_platform": None},
    ) as response:
        assert response.status_code == 200
        events = [json.loads(line) for line in response.iter_lines() if line]

    result_events = [event for event in events if event["event"] == "result"]
    batch_result = next(event for event in events if event["event"] == "batch_result")
    report_folders = {event["report_folder_url"] for event in result_events}

    assert len(result_events) == 2
    assert len(report_folders) == 1
    assert batch_result["report_folder_url"] in report_folders
    assert batch_result["summary_url"].endswith("/summary.json")


def test_import_endpoint_returns_project_report_rows(tmp_path: Path, monkeypatch) -> None:
    from app import database
    from app.main import create_app

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    database.engine = engine
    database.SessionLocal.configure(bind=engine)
    Base.metadata.create_all(bind=engine)

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)

    client = TestClient(create_app())
    with workbook_path.open("rb") as file:
        response = client.post(
            "/api/test-plans/import",
            data={"project_name": "AI学区顾问搜索交互"},
            files={"file": ("cases.xlsx", file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["name"] == "AI学区顾问搜索交互"
    assert payload["total_cases"] == 2
    assert payload["cases"][0]["sequence"] == 1
    assert payload["cases"][0]["case_name"] == "验证点击入口跳转"
    assert payload["cases"][0]["run_count"] == 0
    assert payload["cases"][0]["latest_result"] == "pending"


def test_create_case_endpoint_appends_manual_case_and_can_run(tmp_path: Path, monkeypatch) -> None:
    from app import database
    from app.main import create_app

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    database.engine = engine
    database.SessionLocal.configure(bind=engine)
    Base.metadata.create_all(bind=engine)

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)

    client = TestClient(create_app())
    with workbook_path.open("rb") as file:
        import_response = client.post(
            "/api/test-plans/import",
            data={"project_name": "AI学区顾问搜索交互"},
            files={"file": ("cases.xlsx", file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    plan_id = import_response.json()["id"]
    create_response = client.post(
        f"/api/test-plans/{plan_id}/cases",
        json={
            "case_name": "手工验证搜索按钮",
            "precondition": "用户停留在搜索页",
            "steps": ["输入学区名称", "点击搜索按钮"],
            "expected_result": "展示搜索结果列表",
            "module": "搜索",
            "priority": "高",
            "target_app": "乐有家测试版",
        },
    )

    assert create_response.status_code == 200
    created = create_response.json()
    assert created["sequence"] == 3
    assert created["case_name"] == "手工验证搜索按钮"
    assert created["steps"] == ["输入学区名称", "点击搜索按钮"]
    assert created["target_app"] == "乐有家测试版"

    plan_response = client.get(f"/api/test-plans/{plan_id}")
    plan_payload = plan_response.json()
    assert plan_payload["total_cases"] == 3
    assert plan_payload["cases"][-1]["id"] == created["id"]

    run_response = client.post(f"/api/test-plans/cases/{created['id']}/run", json={"device_udid": None})

    assert run_response.status_code == 200
    assert run_response.json()["case_id"] == created["id"]
    assert "未选择设备" in run_response.json()["result_note"]

    report_response = client.get(f"/api/test-plans/{plan_id}/report")
    assert report_response.status_code == 200
    report_payload = report_response.json()
    assert report_payload["cases"][-1]["target_app"] == "乐有家测试版"
    assert report_payload["cases"][-1]["latest_error_category"] == "no_device"

    plan_export_response = client.get(f"/api/test-plans/{plan_id}/export", params={"format": "json"})
    assert plan_export_response.status_code == 200
    plan_export_payload = plan_export_response.json()
    assert plan_export_payload["filename"] == f"test_report_{plan_id}.json"
    assert plan_export_payload["report"]["summary"]["failed"] == 1

    execution_id = run_response.json()["id"]
    execution_detail_response = client.get(f"/api/test-plans/executions/{execution_id}/detail")
    assert execution_detail_response.status_code == 200
    assert execution_detail_response.json()["error_category"] == "no_device"

    execution_export_response = client.get(
        f"/api/test-plans/executions/{execution_id}/export",
        params={"format": "json"},
    )
    assert execution_export_response.status_code == 200
    execution_export_payload = execution_export_response.json()
    assert execution_export_payload["filename"] == f"execution_report_{execution_id}.json"
    assert execution_export_payload["report"]["error_category"] == "no_device"

    execution_html_response = client.get(
        f"/api/test-plans/executions/{execution_id}/export",
        params={"format": "html"},
    )
    assert execution_html_response.status_code == 200
    assert "单用例执行报告" in execution_html_response.json()["html"]


def test_import_service_returns_detailed_skip_reasons(tmp_path: Path) -> None:
    """测试导入服务返回详细的跳过原因"""
    from app.services.test_plan_import_service import TestPlanImportService

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "*所属系统", "*所属模块", "*用例名称", "*前置条件",
            "*用例步骤", "*预期结果", "*关联需求ID", "用例类型", "用例等级",
        ]
    )
    # 说明行
    sheet.append(["说明行", "说明行", "说明行", "说明行", "说明行", "说明行", "说明行", "说明行", "说明行"])
    # 正常用例
    sheet.append(["系统A", "模块A", "正常用例", "前置", "1.步骤1", "预期", "REQ001", "功能", "高"])
    # 缺少步骤
    sheet.append(["系统A", "模块A", "缺少步骤用例", "前置", "", "预期", "REQ002", "功能", "高"])
    # 缺少预期结果
    sheet.append(["系统A", "模块A", "缺少预期用例", "前置", "1.步骤1", "", "REQ003", "功能", "高"])
    # 空行（会被静默跳过）
    sheet.append(["", "", "", "", "", "", "", "", ""])
    # 另一个正常用例
    sheet.append(["系统B", "模块B", "另一个正常用例", "前置", "1.步骤A\n2.步骤B", "预期结果", "REQ004", "功能", "中"])

    workbook_path = tmp_path / "mixed_cases.xlsx"
    workbook.save(workbook_path)
    db = make_session()

    result = TestPlanImportService(db).import_excel(workbook_path, project_name="混合用例测试")

    assert result.imported_count == 2
    assert result.skipped_count == 3  # 说明行 + 缺少步骤 + 缺少预期结果
    assert len(result.skipped_rows) == 3

    # 检查跳过原因
    reasons = [row.reason for row in result.skipped_rows]
    assert any("说明行" in r for r in reasons)
    assert any("必填字段为空" in r and "用例步骤" in r for r in reasons)
    assert any("必填字段为空" in r and "预期结果" in r for r in reasons)

    # 检查摘要
    summary = result.to_summary()
    assert summary["imported"] == 2
    assert summary["skipped"] == 3


def test_import_service_detects_duplicates(tmp_path: Path) -> None:
    """测试导入服务检测重复用例"""
    from app.services.test_plan_import_service import TestPlanImportService

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "*所属系统", "*所属模块", "*用例名称", "*前置条件",
            "*用例步骤", "*预期结果", "*关联需求ID", "用例类型", "用例等级",
        ]
    )
    sheet.append(["系统A", "模块A", "重复用例", "前置", "1.步骤1", "预期", "REQ001", "功能", "高"])
    sheet.append(["系统A", "模块A", "重复用例", "前置", "1.步骤2", "预期", "REQ002", "功能", "高"])  # 同名
    sheet.append(["系统B", "模块B", "唯一用例", "前置", "1.步骤3", "预期", "REQ003", "功能", "高"])

    workbook_path = tmp_path / "duplicate_cases.xlsx"
    workbook.save(workbook_path)
    db = make_session()

    result = TestPlanImportService(db).import_excel(
        workbook_path,
        project_name="重复检测测试",
        check_duplicates=True,
    )

    assert result.imported_count == 2
    assert result.duplicate_count == 1
    assert any("重复" in row.reason for row in result.skipped_rows)


def test_import_service_validates_data(tmp_path: Path) -> None:
    """测试导入服务数据校验"""
    from app.services.test_plan_import_service import TestPlanImportService

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "*所属系统", "*所属模块", "*用例名称", "*前置条件",
            "*用例步骤", "*预期结果", "*关联需求ID", "用例类型", "用例等级",
        ]
    )
    # 正常用例
    sheet.append(["系统A", "模块A", "正常用例", "前置", "1.步骤1", "预期", "REQ001", "功能", "高"])
    # 用例名称超长（超过500字符）
    long_name = "超长用例名称" * 100
    sheet.append(["系统A", "模块A", long_name, "前置", "1.步骤1", "预期", "REQ002", "功能", "高"])
    # 步骤过多（超过50条）
    too_many_steps = "\n".join(f"{i}.步骤{i}" for i in range(1, 60))
    sheet.append(["系统A", "模块A", "步骤过多用例", "前置", too_many_steps, "预期", "REQ003", "功能", "高"])

    workbook_path = tmp_path / "validation_cases.xlsx"
    workbook.save(workbook_path)
    db = make_session()

    result = TestPlanImportService(db).import_excel(workbook_path, project_name="数据校验测试")

    # 超长名称会被截断，步骤过多的会被跳过
    assert result.imported_count >= 1
    # 检查是否有数据校验失败的跳过记录
    validation_errors = [row for row in result.skipped_rows if "数据校验失败" in row.reason]
    assert len(validation_errors) >= 1


def test_import_endpoint_returns_import_summary(tmp_path: Path) -> None:
    """测试导入接口返回导入摘要"""
    from app import database
    from app.main import create_app

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    database.engine = engine
    database.SessionLocal.configure(bind=engine)
    Base.metadata.create_all(bind=engine)

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)

    client = TestClient(create_app())
    with workbook_path.open("rb") as file:
        response = client.post(
            "/api/test-plans/import",
            data={"project_name": "导入摘要测试"},
            files={"file": ("cases.xlsx", file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert response.status_code == 200
    payload = response.json()

    # 检查返回了导入摘要
    assert "import_summary" in payload
    assert "imported" in payload["import_summary"]
    assert "skipped" in payload["import_summary"]
    assert payload["import_summary"]["imported"] == 2

    # 检查返回了跳过行详情
    assert "skipped_rows" in payload
    assert isinstance(payload["skipped_rows"], list)


def test_batch_update_cases_endpoint_updates_and_clears_fields(tmp_path: Path) -> None:
    from app import database
    from app.main import create_app

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    database.engine = engine
    database.SessionLocal.configure(bind=engine)
    Base.metadata.create_all(bind=engine)

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)

    client = TestClient(create_app())
    with workbook_path.open("rb") as file:
        import_response = client.post(
            "/api/test-plans/import",
            data={"project_name": "批量编辑测试"},
            files={"file": ("cases.xlsx", file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    payload = import_response.json()
    plan_id = payload["id"]
    case_ids = [payload["cases"][0]["id"], payload["cases"][1]["id"]]

    update_response = client.post(
        f"/api/test-plans/{plan_id}/cases/batch-update",
        json={
            "case_ids": case_ids,
            "system_name": "  乐有家  ",
            "module": "",
            "target_app": "乐有家测试版",
            "test_module": "AI找房",
        },
    )

    assert update_response.status_code == 200
    assert update_response.json()["updated"] == 2

    plan_response = client.get(f"/api/test-plans/{plan_id}")
    assert plan_response.status_code == 200
    cases = plan_response.json()["cases"]
    assert {case["system_name"] for case in cases} == {"乐有家"}
    assert {case["module"] for case in cases} == {None}
    assert {case["target_app"] for case in cases} == {"乐有家测试版"}
    assert {case["test_module"] for case in cases} == {"AI找房"}


def test_batch_update_cases_endpoint_rejects_invalid_target_app(tmp_path: Path) -> None:
    from app import database
    from app.main import create_app

    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    database.engine = engine
    database.SessionLocal.configure(bind=engine)
    Base.metadata.create_all(bind=engine)

    workbook_path = tmp_path / "cases.xlsx"
    build_workbook(workbook_path)

    client = TestClient(create_app())
    with workbook_path.open("rb") as file:
        import_response = client.post(
            "/api/test-plans/import",
            data={"project_name": "批量编辑测试"},
            files={"file": ("cases.xlsx", file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    payload = import_response.json()
    plan_id = payload["id"]
    case_id = payload["cases"][0]["id"]

    update_response = client.post(
        f"/api/test-plans/{plan_id}/cases/batch-update",
        json={
            "case_ids": [case_id],
            "target_app": "无效应用",
        },
    )

    assert update_response.status_code == 422
